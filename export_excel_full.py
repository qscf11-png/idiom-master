import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_all_to_excel():
    print("正在讀取全量整合數據...")
    try:
        with open("master_vocab_full.json", "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"讀取失敗: {e}")
        return
        
    print(f"總計讀取 {len(data)} 個單字。")
    
    # Prepare list for DataFrame
    rows = []
    # 按照 Level 排序，同 Level 按照單字字母排序
    sorted_words = sorted(data.items(), key=lambda x: (x[1].get("level", 9), x[0]))
    
    for word, details in sorted_words:
        rows.append({
            "Level": f"Level {details.get('level', 'Unknown')}",
            "Vocabulary": word,
            "IPA": details.get("ipa", ""),
            "Part of Speech / Definition": details.get("def", ""),
            "Inflection": details.get("trans", ""),
            "Collocation": details.get("col", ""),
            "Example Sentence": details.get("ex", "")
        })
        
    df = pd.DataFrame(rows)
    
    output_path = "Full_Vocabulary_Levels_1-6.xlsx"
    
    print(f"正在生成 Excel: {output_path}...")
    
    # Use ExcelWriter for styling
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Vocabulary List')
        
        workbook = writer.book
        worksheet = writer.sheets['Vocabulary List']
        
        # Styling
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Apply styles to header
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
        # Set column widths
        worksheet.column_dimensions['A'].width = 12 # Level
        worksheet.column_dimensions['B'].width = 20 # Vocabulary
        worksheet.column_dimensions['C'].width = 15 # IPA
        worksheet.column_dimensions['D'].width = 40 # Def
        worksheet.column_dimensions['E'].width = 25 # Inflection
        worksheet.column_dimensions['F'].width = 35 # Collocation
        worksheet.column_dimensions['G'].width = 60 # Example
        
        # Apply borders and alignment to data cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border
                # Col 1, 2, 3 center aligned
                if cell.column <= 3:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
                    
        # Freeze top row
        worksheet.freeze_panes = "A2"

    print(f"\nExcel 匯出完成！最終單字數: {len(df)}")
    print(f"檔案路徑: {os.path.abspath(output_path)}")

if __name__ == "__main__":
    import os
    export_all_to_excel()
