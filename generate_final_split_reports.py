import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def generate_split_reports_v2():
    input_path = "master_vocab_full.json"
    if not os.path.exists(input_path):
        print(f"Error: {input_path} not found.")
        return

    print(f"正在讀取資料集: {input_path}...")
    with open(input_path, "r", encoding="utf-8") as f:
        master_data = json.load(f)

    # 按照 Level 分組
    levels_data = {i: [] for i in range(1, 7)}
    for word, details in master_data.items():
        lvl = details.get("level", 0)
        if 1 <= lvl <= 6:
            levels_data[lvl].append((word, details))

    for lvl, items in levels_data.items():
        if not items:
            continue

        # 字母排序單字
        items.sort(key=lambda x: x[0])

        wb = Workbook()
        ws = wb.active
        ws.title = "Vocabulary"

        # 使用簡化標題，避免 SheetJS 在處理特殊字元 (+) 或長標題時出錯
        headers = ["Vocabulary", "Content", "POS"]
        ws.append(headers)

        # 設定樣式
        header_font = Font(name='Calibri', bold=True, size=12)
        data_font = Font(name='Calibri', size=11)
        wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment

        # 寫入資料
        for word, details in items:
            # 建立內容字串
            # 為了讓 HTML 工具能讀到，我們確保每項資訊都分行，且不使用會被過濾的 【 】
            content_parts = []
            if details.get("def"): content_parts.append(details['def']) # 解釋放第一行
            if details.get("ipa"): content_parts.append(f"IPA: {details['ipa']}")
            if details.get("trans"): content_parts.append(f"Trans: {details['trans']}")
            if details.get("col"): content_parts.append(f"Col: {details['col']}")
            if details.get("ex"): content_parts.append(f"Ex: {details['ex']}")
            
            combined_content = "\n".join(content_parts)
            
            row_data = [word, combined_content, details.get("pos", "")]
            ws.append(row_data)

            # 套用內容樣式 (目前最後一行)
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=1).alignment = center_alignment
            ws.cell(row=row_idx, column=2).alignment = wrap_alignment
            ws.cell(row=row_idx, column=3).alignment = center_alignment
            
            for col_idx in range(1, 4):
                ws.cell(row=row_idx, column=col_idx).font = data_font

        # 設定寬度
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 15
        
        ws.freeze_panes = "A2"

        output_name = f"高中英文參考詞彙表_Level_{lvl}.xlsx"
        wb.save(output_name)
        print(f"Level {lvl} 報表產出完成: {output_name} ({len(items)} 筆)")

if __name__ == "__main__":
    generate_split_reports_v2()
